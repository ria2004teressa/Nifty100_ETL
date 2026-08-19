# ============================================================
# NIFTY 100 SCREENER
# SPRINT 3 - DAY 17
# COMPOSITE SCORE + EXCEL EXPORT
# ============================================================

from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Project root
PROJECT_ROOT = Path(__file__).resolve().parents[2]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.screener.engine import apply_preset


# ============================================================
# PATHS
# ============================================================

OUTPUT_DIR = PROJECT_ROOT / "output"

OUTPUT_FILE = (
    OUTPUT_DIR /
    "screener_output.xlsx"
)


# ============================================================
# PRESETS
# ============================================================

PRESETS = {
    "quality_compounder": "Quality Compounder",
    "value_pick": "Value Pick",
    "growth_accelerator": "Growth Accelerator",
    "dividend_champion": "Dividend Champion",
    "debt_free_blue_chip": "Debt-Free Blue Chip",
    "turnaround_watch": "Turnaround Watch",
}


# ============================================================
# COLUMNS TO DISPLAY
# ============================================================

KPI_COLUMNS = [
    "company_id",
    "year",
    "company_name",
    "broad_sector",

    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    "debt_to_equity",
    "interest_coverage",

    "free_cash_flow_cr",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "asset_turnover",

    "sales",
    "net_profit",

    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",

    "dividend_payout_ratio_pct",

    "market_cap_cr",

    "composite_quality_score",
]


# ============================================================
# CREATE OUTPUT DIRECTORY
# ============================================================

def create_output_directory():

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True
    )


# ============================================================
# PREPARE RESULT
# ============================================================

def prepare_result(
    result
):

    if result.empty:
        return result

    available = [
        column
        for column in KPI_COLUMNS
        if column in result.columns
    ]

    result = result[
        available
    ].copy()

    # Sort by composite score
    if "composite_quality_score" in result.columns:

        result = result.sort_values(
            "composite_quality_score",
            ascending=False
        )

    return result.reset_index(
        drop=True
    )


# ============================================================
# EXPORT ALL PRESETS
# ============================================================

def export_presets():

    create_output_directory()

    print()
    print("=" * 70)
    print("GENERATING SCREENER OUTPUT")
    print("=" * 70)

    with __import__("pandas").ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for preset_key, preset_name in PRESETS.items():

            print()
            print(
                f"Processing: {preset_name}"
            )

            result = apply_preset(
                preset_key
            )

            result = prepare_result(
                result
            )

            # Excel sheet names have a 31-character limit
            sheet_name = preset_name[:31]

            if result.empty:

                # Create an empty sheet with a message
                import pandas as pd

                empty_df = pd.DataFrame(
                    {
                        "Message": [
                            "No companies matched this preset."
                        ]
                    }
                )

                empty_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )

                print(
                    f"  Results: 0"
                )

            else:

                result.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )

                print(
                    f"  Results: {len(result)}"
                )

    print()
    print(
        f"Excel file created:"
    )

    print(
        OUTPUT_FILE
    )

    return OUTPUT_FILE


# ============================================================
# FORMAT EXCEL
# ============================================================

def format_excel():

    if not OUTPUT_FILE.exists():

        raise FileNotFoundError(
            f"Excel file not found: {OUTPUT_FILE}"
        )

    workbook = load_workbook(
        OUTPUT_FILE
    )


    # --------------------------------------------------------
    # Fills
    # --------------------------------------------------------

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )


    # --------------------------------------------------------
    # Process each sheet
    # --------------------------------------------------------

    for worksheet in workbook.worksheets:

        # Header formatting
        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.fill = header_fill


        # Freeze header
        worksheet.freeze_panes = "A2"


        # Auto width
        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:

                    length = len(
                        str(cell.value)
                    )

                    if length > max_length:

                        max_length = length


            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                30
            )


        # ----------------------------------------------------
        # Find columns
        # ----------------------------------------------------

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }


        # ----------------------------------------------------
        # Composite score formatting
        # ----------------------------------------------------

        if "composite_quality_score" in headers:

            column_number = headers[
                "composite_quality_score"
            ]

            for row in range(
                2,
                worksheet.max_row + 1
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column_number
                )

                if isinstance(
                    cell.value,
                    (int, float)
                ):

                    if cell.value >= 75:

                        cell.fill = green_fill

                    elif cell.value < 50:

                        cell.fill = red_fill


        # ----------------------------------------------------
        # Basic KPI threshold formatting
        # ----------------------------------------------------

        threshold_rules = {

            "return_on_equity_pct": (
                15,
                "min"
            ),

            "debt_to_equity": (
                1,
                "max"
            ),

            "free_cash_flow_cr": (
                0,
                "min"
            ),

            "revenue_cagr_5yr": (
                10,
                "min"
            ),

            "pat_cagr_5yr": (
                20,
                "min"
            ),

            "dividend_yield_pct": (
                2,
                "min"
            ),

        }


        for metric, rule in threshold_rules.items():

            if metric not in headers:

                continue


            threshold, rule_type = rule

            column_number = headers[
                metric
            ]


            for row in range(
                2,
                worksheet.max_row + 1
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column_number
                )


                if not isinstance(
                    cell.value,
                    (int, float)
                ):

                    continue


                if rule_type == "min":

                    if cell.value >= threshold:

                        cell.fill = green_fill

                    else:

                        cell.fill = red_fill


                elif rule_type == "max":

                    if cell.value <= threshold:

                        cell.fill = green_fill

                    else:

                        cell.fill = red_fill


    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    workbook.save(
        OUTPUT_FILE
    )

    print()
    print(
        "Excel formatting completed."
    )


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    print()
    print("=" * 70)
    print("NIFTY 100 - DAY 17")
    print("COMPOSITE SCORE + EXCEL EXPORT")
    print("=" * 70)

    export_presets()

    format_excel()

    print()
    print("=" * 70)
    print("DAY 17 COMPLETED")
    print("=" * 70)